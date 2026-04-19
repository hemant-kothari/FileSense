"""
AI Engine — Groq API calls, file content extraction, sensitive data detection.

v2 adds:
  • generate_tags()       — 3–5 semantic tags in the same API call as describe_file()
  • generate_narrative()  — 60–200 word evolution narrative (separate call)
  • Shannon entropy check in detect_sensitive()
  • Expanded sensitive patterns (JWT, Stripe, Twilio, SendGrid, PII)
"""

import base64
import io
import json
import math
import os
import re
from pathlib import Path
from typing import Optional

import requests

# Sensitive data patterns
# Each tuple: (regex_pattern, human_label)
# Patterns are checked in order; first match per label wins.

SENSITIVE_PATTERNS = [
    # ── API / secret keys ─────────────────────────────────────────────────────
    (r'(?i)(api[_\-\s]?key|apikey)\s*[=:]\s*["\']?([A-Za-z0-9\-_]{20,})',
     "API key"),
    (r'(?i)(secret[_\-\s]?key|client_secret)\s*[=:]\s*["\']?([A-Za-z0-9\-_]{20,})',
     "secret key"),
    (r'(?i)(password|passwd|pwd)\s*[=:]\s*["\']?\S+',
     "password"),
    (r'-----BEGIN (?:RSA |EC |OPENSSH |DSA )?PRIVATE KEY-----',
     "private key"),
    (r'(?i)(aws_access_key_id|aws_secret_access_key)\s*[=:]\s*["\']?[A-Za-z0-9/+]{20,}',
     "AWS credential"),

    # Provider-specific tokens
    (r'sk-[A-Za-z0-9]{48}',                                         "OpenAI API key"),
    (r'AIza[0-9A-Za-z\-_]{35}',                                     "Google API key"),
    (r'(?i)bearer\s+[A-Za-z0-9\-._~+/]+=*',                        "bearer token"),
    (r'ghp_[A-Za-z0-9]{36}',                                        "GitHub token"),
    (r'gho_[A-Za-z0-9]{36}',                                        "GitHub OAuth token"),
    (r'ghs_[A-Za-z0-9]{36}',                                        "GitHub app token"),
    (r'xox[baprs]-[A-Za-z0-9\-]{10,}',                              "Slack token"),
    (r'gsk_[A-Za-z0-9]{52}',                                        "Groq API key"),
    (r'sk_live_[A-Za-z0-9]{24}',                                    "Stripe live secret key"),
    (r'pk_live_[A-Za-z0-9]{24}',                                    "Stripe live publishable key"),
    (r'rk_live_[A-Za-z0-9]{24}',                                    "Stripe restricted key"),
    (r'AC[a-f0-9]{32}',                                             "Twilio Account SID"),
    (r'SK[a-f0-9]{32}',                                             "Twilio API key SID"),
    (r'SG\.[A-Za-z0-9_-]{22}\.[A-Za-z0-9_-]{43}',                  "SendGrid API key"),
    (r'EAAA[A-Za-z0-9]{80,}',                                       "Facebook access token"),
    (r'ya29\.[A-Za-z0-9_\-]{60,}',                                  "Google OAuth token"),

    # JWTs
    (r'eyJ[A-Za-z0-9\-_]{20,}\.[A-Za-z0-9\-_]{20,}\.[A-Za-z0-9\-_]{20,}',
     "JWT token"),

    # Connection strings / URLs 
    (r'(?i)database_url\s*[=:]\s*["\']?\S+',                        "database URL"),
    (r'(?i)connection.?string\s*[=:]\s*["\']?[^\n"\']+',            "connection string"),
    (r'(?i)(mongodb|postgresql|mysql|redis|amqp)://[^\s"\']+',       "database connection"),

    #  PII
    (r'\b\d{3}-\d{2}-\d{4}\b',                                      "SSN (PII)"),
    (r'\b(?:4[0-9]{12}(?:[0-9]{3})?|5[1-5][0-9]{14}|3[47][0-9]{13})\b',
     "credit card number (PII)"),
]

# Words the model echoes back when it misreads the prompt as a template
_BAD_PHRASES = [
    "short description", "less than 20 words", "max 20 words",
    "less than 60 words", "max 60 words", "insert description",
    "placeholder", "your description here", "description goes here",
    "write a", "provide a", "example description",
]

# File type groups

TEXT_EXTS = {
    ".py", ".js", ".ts", ".jsx", ".tsx", ".cpp", ".c", ".h", ".hpp",
    ".java", ".go", ".rs", ".rb", ".php", ".swift", ".kt", ".cs",
    ".html", ".css", ".scss", ".sass", ".less", ".md", ".txt",
    ".yaml", ".yml", ".toml", ".ini", ".cfg", ".sh", ".bash",
    ".zsh", ".ps1", ".sql", ".graphql", ".xml", ".env",
    ".gitignore", ".dockerfile", ".makefile", ".r", ".lua",
}
DATA_EXTS  = {".csv", ".tsv", ".json", ".jsonl"}
EXCEL_EXTS = {".xlsx", ".xls"}
DOC_EXTS   = {".pdf", ".docx", ".doc", ".odt"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff", ".svg"}

# Narrative word-count targets by content mode
_NARRATIVE_TARGETS: dict[str, tuple[int, int]] = {
    "text":      (100, 200),   # code / scripts / config
    "document":  (80,  150),   # pdf / docx / markdown
    "data":      (60,  100),   # csv / json / xlsx
    "image":     (60,   80),   # images
    "too_large": (40,   60),   # over size limit
    "binary":    (40,   60),
    "error":     (20,   40),
}


# Sensitive-data detection

def _shannon_entropy(s: str) -> float:
    """Calculate Shannon entropy of a string (bits per character)."""
    if not s:
        return 0.0
    freq: dict[str, int] = {}
    for c in s:
        freq[c] = freq.get(c, 0) + 1
    n = len(s)
    return -sum((f / n) * math.log2(f / n) for f in freq.values())


def detect_sensitive(content: str) -> tuple[bool, list[str]]:
    """
    Scan content for sensitive data patterns.
    Returns (has_sensitive: bool, labels: list[str]).
    """
    found: list[str] = []

    # 1. Regex pattern matching
    for pattern, label in SENSITIVE_PATTERNS:
        if re.search(pattern, content) and label not in found:
            found.append(label)

    # 2. Shannon entropy detector — catches obfuscated / custom-named secrets
    #    Any token ≥ 20 characters with entropy > 4.5 bits/char is suspicious.
    #    (High-entropy → random-looking → likely a key/token.)
    entropy_label = "high-entropy string (possible secret)"
    if entropy_label not in found:
        for token in re.findall(r'[A-Za-z0-9+/=_\-]{20,}', content):
            if _shannon_entropy(token) > 4.5:
                found.append(entropy_label)
                break

    return bool(found), found


def _looks_like_template(text: str) -> bool:
    low = text.lower()
    return any(p in low for p in _BAD_PHRASES)


# File content extraction

def extract_content(file_path: str, cfg: dict) -> tuple[str, bool, list[str], str]:
    """
    Extract text/data content from a file for AI ingestion.
    Returns: (content_str, is_sensitive, sensitive_types, mode)
    mode ∈ {"text", "data", "document", "image", "binary", "too_large", "error"}
    """
    path   = Path(file_path)
    ext    = path.suffix.lower()
    max_mb = cfg.get("max_file_size_mb", 10)

    try:
        size_mb = os.path.getsize(file_path) / 1_048_576
    except OSError:
        return "[Cannot read file]", False, [], "error"

    if size_mb > max_mb:
        return f"[Large file: {size_mb:.1f} MB, type: {ext or 'unknown'}]", False, [], "too_large"

    # Plain text / code
    if ext in TEXT_EXTS or (ext == "" and size_mb < 1):
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                raw = f.read()
            head = cfg.get("max_text_chars", 3000)
            tail = cfg.get("max_tail_chars", 500)
            content = (
                raw[:head] + "\n...[truncated]...\n" + raw[-tail:]
                if len(raw) > head + tail else raw
            )
            sensitive, stypes = detect_sensitive(raw)
            return content, sensitive, stypes, "text"
        except Exception:
            pass

    # CSV / TSV
    if ext in {".csv", ".tsv"}:
        try:
            import csv
            sep  = "\t" if ext == ".tsv" else ","
            rows = []
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                for i, row in enumerate(csv.reader(f, delimiter=sep)):
                    rows.append(sep.join(str(v) for v in row))
                    if i >= cfg.get("max_csv_rows", 20):
                        break
            total   = sum(1 for _ in open(file_path, encoding="utf-8", errors="replace"))
            snippet = f"[CSV — {total} rows total, showing first {len(rows)}]\n" + "\n".join(rows)
            sensitive, stypes = detect_sensitive(snippet)
            return snippet, sensitive, stypes, "data"
        except Exception as e:
            return f"[CSV, read error: {e}]", False, [], "data"

    # JSON / JSONL
    if ext in {".json", ".jsonl"}:
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                raw = f.read()
            content   = raw[:cfg.get("max_text_chars", 3000)]
            sensitive, stypes = detect_sensitive(raw)
            return content, sensitive, stypes, "data"
        except Exception as e:
            return f"[JSON, read error: {e}]", False, [], "data"

    # Excel
    if ext in EXCEL_EXTS:
        try:
            import openpyxl
            wb    = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            parts = []
            for sname in list(wb.sheetnames)[:3]:
                ws = wb[sname]
                parts.append(f"[Sheet: {sname}]")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    parts.append("\t".join("" if v is None else str(v) for v in row))
                    if i >= cfg.get("max_csv_rows", 20):
                        parts.append("...[more rows]...")
                        break
            snippet   = "\n".join(parts)
            sensitive, stypes = detect_sensitive(snippet)
            return snippet, sensitive, stypes, "data"
        except Exception as e:
            return f"[Excel, read error: {e}]", False, [], "data"

    #  PDF 
    if ext == ".pdf":
        try:
            import pypdf
            reader = pypdf.PdfReader(file_path)
            cap    = cfg.get("max_text_chars", 3000)
            texts  = []
            for page in reader.pages[:10]:
                texts.append(page.extract_text() or "")
                if sum(len(t) for t in texts) >= cap:
                    break
            raw = "\n".join(texts)[:cap]
            sensitive, stypes = detect_sensitive(raw)
            return raw, sensitive, stypes, "document"
        except Exception as e:
            return f"[PDF, read error: {e}]", False, [], "document"

    # Word
    if ext in {".docx", ".doc"}:
        try:
            from docx import Document
            doc = Document(file_path)
            raw = "\n".join(p.text for p in doc.paragraphs)
            cap = cfg.get("max_text_chars", 3000)
            sensitive, stypes = detect_sensitive(raw)
            return raw[:cap], sensitive, stypes, "document"
        except Exception as e:
            return f"[Word doc, read error: {e}]", False, [], "document"

    # Images
    if ext in IMAGE_EXTS:
        try:
            from PIL import Image as PILImage
            img = PILImage.open(file_path).convert("RGB")
            img.thumbnail((512, 512))
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            b64 = base64.b64encode(buf.getvalue()).decode()
            return f"__IMG__image/jpeg|{b64}", False, [], "image"
        except Exception as e:
            return f"[Image {ext}, preview error: {e}]", False, [], "image"

    return f"[Binary file, type: {ext or 'unknown'}]", False, [], "binary"

# AI Engine

class AIEngine:
    def __init__(self, config: dict):
        self.config    = config
        self.api_key   = config.get("groq_api_key", "")
        self.model     = config.get("groq_model", "llama-3.1-8b-instant")
        self.vis_model = config.get("groq_vision_model",
                                    "meta-llama/llama-4-scout-17b-16e-instruct")

    def is_configured(self) -> bool:
        return bool(self.api_key.strip())

    def update_config(self, config: dict):
        self.config    = config
        self.api_key   = config.get("groq_api_key", "")
        self.model     = config.get("groq_model", "llama-3.1-8b-instant")
        self.vis_model = config.get("groq_vision_model",
                                    "meta-llama/llama-4-scout-17b-16e-instruct")

    # Low-level Groq call

    def _call(
        self,
        messages: list,
        model: str | None = None,
        system: str | None = None,
        max_tokens: int = 450,
    ) -> Optional[str]:
        if not self.api_key:
            return None
        full = []
        if system:
            full.append({"role": "system", "content": system})
        full.extend(messages)
        payload = {
            "model":       model or self.model,
            "messages":    full,
            "max_tokens":  max_tokens,
            "temperature": 0.2,
        }
        try:
            r = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type":  "application/json",
                },
                json=payload,
                timeout=30,
            )
            print(f"[groq] status: {r.status_code}")
            r.raise_for_status()
            return r.json()["choices"][0]["message"]["content"].strip()
        except Exception as e:
            print(f"[groq] error: {e}")
            return None

    # File description + tags (single API call)

    def describe_file(self, file_path: str) -> dict:
        """
        Generate short description, long description, and semantic tags
        for a file — all in one Groq API call.

        Returns dict with keys:
          short, long, tags, sensitive_detected, sensitive_types
        """
        path    = Path(file_path)
        content, sensitive, stypes, mode = extract_content(file_path, self.config)

        if mode == "image" and content.startswith("__IMG__"):
            return self._describe_image(path.name, content, sensitive, stypes)

        sens_note = (
            f"\nThis file contains sensitive data ({', '.join(stypes)}). "
            "Do NOT reveal any actual values — only mention the type of sensitive data present."
        ) if sensitive else ""

        system = (
            "You are a file annotation tool. Write factual, specific descriptions "
            "and extract semantic topic tags based on file content. "
            "Respond with a single raw JSON object only — "
            "no markdown fences, no explanation, no preamble."
        )
        prompt = (
            f"Analyse this file and provide descriptions plus semantic tags.\n\n"
            f"Filename: {path.name}\n"
            f"Content type: {mode}{sens_note}\n\n"
            f"Content (may be truncated):\n---\n{content}\n---\n\n"
            f"Respond with exactly:\n"
            f'{{"short": "one factual sentence under 20 words", '
            f'"long": "two or three factual sentences under 60 words total", '
            f'"tags": ["3 to 5 lowercase topic tags, 1-2 words each"]}}'
        )

        result = self._parse(
            self._call([{"role": "user", "content": prompt}], system=system),
            sensitive, stypes,
        )

        # One retry if the model echoed back template text
        if _looks_like_template(result["short"]) or _looks_like_template(result["long"]):
            retry = (
                f"Describe the file '{path.name}' based on this content preview:\n"
                f"{content[:800]}\n\n"
                f'Reply with only: {{"short": "...", "long": "...", "tags": ["..."]}}'
            )
            result = self._parse(
                self._call([{"role": "user", "content": retry}], system=system),
                sensitive, stypes,
            )

        return result

    def _describe_image(self, name: str, content: str, sensitive: bool, stypes: list) -> dict:
        _, rest      = content.split("__IMG__", 1)
        media_type, b64 = rest.split("|", 1)
        system   = (
            "You are a file annotation tool. "
            "Respond with a single raw JSON object only. No markdown, no preamble."
        )
        messages = [{
            "role": "user",
            "content": [
                {"type": "image_url",
                 "image_url": {"url": f"data:{media_type};base64,{b64}"}},
                {"type": "text",
                 "text": (
                     f"Describe the image file '{name}' and suggest tags.\n"
                     'Reply with only: {"short": "one sentence under 20 words", '
                     '"long": "two or three sentences under 60 words", '
                     '"tags": ["2 to 4 descriptive tags"]}'
                 )},
            ],
        }]
        return self._parse(
            self._call(messages, model=self.vis_model, system=system),
            sensitive, stypes,
        )

    # Folder description 

    def describe_folder(self, folder_path: str, file_descs: dict) -> dict:
        name = Path(folder_path).name

        real = {
            fn: v for fn, v in file_descs.items()
            if v.get("short_desc") and not _looks_like_template(v.get("short_desc", ""))
        }
        any_sensitive = any(v.get("sensitive_detected") for v in file_descs.values())

        if not real:
            names   = list(file_descs.keys())
            if not names:
                return {
                    "short": f"Empty folder: {name}",
                    "long":  f"The folder '{name}' contains no files.",
                    "tags":  [],
                    "sensitive_detected": False,
                }
            preview = ", ".join(names[:4])
            return {
                "short": f"{name} — {len(names)} file(s): {', '.join(names[:3])}",
                "long":  f"Contains {len(names)} file(s) including {preview}.",
                "tags":  [],
                "sensitive_detected": any_sensitive,
            }

        lines   = [f"- {fn}: {v.get('short_desc', '')}" for fn, v in list(real.items())[:20]]
        summary = "\n".join(lines)
        sens    = (
            " Note: some files contain sensitive data — mention this briefly."
            if any_sensitive else ""
        )

        system = (
            "You are a folder annotation tool. Summarise what a folder does "
            "based on its file descriptions. Be specific and factual. "
            "Respond with a single raw JSON object only."
        )
        prompt = (
            f"Write a description and tags for the folder '{name}'.{sens}\n\n"
            f"Its files:\n{summary}\n\n"
            f'Reply with only: {{"short": "one factual sentence under 20 words", '
            f'"long": "two or three factual sentences under 60 words total", '
            f'"tags": ["2 to 4 topic tags"]}}'
        )

        raw    = self._call([{"role": "user", "content": prompt}], system=system)
        result = self._parse(raw, any_sensitive, [])

        if _looks_like_template(result["short"]):
            retry  = (
                f"Describe the folder '{name}' which contains:\n{summary[:600]}\n"
                f'Reply with only: {{"short": "...", "long": "...", "tags": ["..."]}}'
            )
            result = self._parse(
                self._call([{"role": "user", "content": retry}], system=system),
                any_sensitive, [],
            )

        result["sensitive_detected"] = any_sensitive
        return result

    # Narrative generation

    def generate_narrative(
        self,
        file_path: str,
        old_narrative: str,
        new_short: str,
        new_long: str,
        history: list[dict],
        mode: str = "text",
    ) -> str:
        """
        Generate or update a 60–200 word evolution narrative for a file.

        • First description  → narrative is a fleshed-out, present-tense summary.
        • Subsequent updates → narrative incorporates history to show evolution.

        Returns plain text (not JSON).
        """
        if not self.api_key:
            return ""

        min_w, max_w = _NARRATIVE_TARGETS.get(mode, (60, 100))
        name         = Path(file_path).name

        if not history and not old_narrative:
            # First time — write a narrative overview from the description alone
            prompt = (
                f"Write a {min_w}–{max_w} word narrative overview for the file '{name}'.\n"
                "Explain what the file does, its purpose, and its role in the project.\n"
                "Use present tense. Be specific and factual.\n\n"
                f"Current description:\n"
                f"  Short:    {new_short}\n"
                f"  Extended: {new_long}\n\n"
                "Write ONLY the narrative text. No labels, no JSON, no bullet points."
            )
        else:
            history_lines = "\n".join(
                f"  • {h.get('recorded_at', '')[:10]} ({h.get('source','ai')}): "
                f"{h.get('short','')}"
                for h in history[:3]
            )
            prompt = (
                f"Update the narrative for the file '{name}' ({min_w}–{max_w} words).\n"
                "Describe what the file does now AND how it has evolved over time.\n"
                "Use present tense for current state, past tense for history.\n"
                "Be specific about meaningful changes; skip trivial ones.\n\n"
                f"Current description:\n"
                f"  Short:    {new_short}\n"
                f"  Extended: {new_long}\n\n"
                + (f"Previous versions:\n{history_lines}\n\n" if history_lines else "")
                + (f"Previous narrative (for context):\n{old_narrative[:400]}\n\n"
                   if old_narrative else "")
                + "Write ONLY the updated narrative text. No labels, no JSON."
            )

        system = (
            "You are a technical writer creating concise file-evolution narratives. "
            "Write exactly what is asked — no preamble, no labels, just the narrative text."
        )
        raw = self._call(
            [{"role": "user", "content": prompt}],
            system=system,
            max_tokens=max_w * 3,  # generous token budget
        )
        return (raw or "").strip()

    # Response parser

    @staticmethod
    def _parse(raw: Optional[str], sensitive: bool, stypes: list) -> dict:
        default = {
            "short":              "Could not generate description.",
            "long":               "AI description generation failed for this item.",
            "tags":               [],
            "sensitive_detected": sensitive,
            "sensitive_types":    stypes,
        }
        if not raw:
            return default
        try:
            clean = re.sub(r"^```[a-z]*\s*", "", raw.strip())
            clean = re.sub(r"```\s*$",        "", clean).strip()
            brace = clean.find("{")
            if brace > 0:
                clean = clean[brace:]
            data  = json.loads(clean)
            short = str(data.get("short", "")).strip()
            long_ = str(data.get("long",  "")).strip()
            tags  = data.get("tags", [])
            if not isinstance(tags, list):
                tags = []
            # Normalise tags: lowercase, strip whitespace, max 5
            tags = [str(t).lower().strip() for t in tags if t][:5]
            if not short:
                return default
            return {
                "short":              short[:200],
                "long":               long_[:600],
                "tags":               tags,
                "sensitive_detected": sensitive,
                "sensitive_types":    stypes,
            }
        except Exception:
            # Fallback: try regex extraction for short + long only
            sm = re.search(r'"short"\s*:\s*"([^"]+)"', raw)
            lm = re.search(r'"long"\s*:\s*"([^"]+)"',  raw)
            if sm:
                return {
                    "short":              sm.group(1)[:200],
                    "long":               lm.group(1)[:600] if lm else "",
                    "tags":               [],
                    "sensitive_detected": sensitive,
                    "sensitive_types":    stypes,
                }
            return default